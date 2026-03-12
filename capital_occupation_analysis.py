import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import datetime

# ================= 配置区域 =================
ap_pattern = "1413-AP-*.xlsx"
ar_pattern = "1413-AR-*.xlsx"

ap_opening_file = "1413-AP-2022-余额.XLSX"
ar_opening_file = "1413-AR-2022-余额.XLSX"

output_data_file = "data.xlsx"
output_report_file = "资金占用日报表_分客户版.xlsx"

annual_rate = 0.04          
threshold_ratio = 0.05      
# ===========================================

# 【新增】获取当前日期，用于计算期末利息
today = datetime.date.today()
print(f"📅 当前日期：{today}")

print("="*80)
print("🚀 开始执行：含期末利息补算版本")
print("="*80)

# --------------------------
# 0. 读取并处理期初余额
# --------------------------
print("\n💰 正在读取 2022 年期初余额文件...")
opening_ap_dict = {}
opening_ar_dict = {}

if os.path.exists(ap_opening_file):
    try:
        df_ap_open = pd.read_excel(ap_opening_file)
        temp_ap = df_ap_open.iloc[:, [3, 11]].copy()
        temp_ap.columns = ['交易对象名称', '期初金额']
        opening_ap_group = temp_ap.groupby('交易对象名称')['期初金额'].sum().to_dict()
        opening_ap_dict.update(opening_ap_group)
        print(f"   ✅ AP 期初文件读取成功，共 {len(opening_ap_dict)} 个供应商")
    except Exception as e:
        print(f"   ❌ AP 期初文件读取失败：{e}")

if os.path.exists(ar_opening_file):
    try:
        df_ar_open = pd.read_excel(ar_opening_file)
        temp_ar = df_ar_open.iloc[:, [5, 13]].copy()
        temp_ar.columns = ['交易对象名称', '期初金额']
        opening_ar_group = temp_ar.groupby('交易对象名称')['期初金额'].sum().to_dict()
        opening_ar_dict.update(opening_ar_group)
        print(f"   ✅ AR 期初文件读取成功，共 {len(opening_ar_dict)} 个客户")
    except Exception as e:
        print(f"   ❌ AR 期初文件读取失败：{e}")

# --------------------------
# 1. 读取并合并流水数据
# --------------------------
print("\n📂 正在扫描流水文件...")
all_data = []

ap_files = [f for f in glob.glob(ap_pattern) if "2022-余额" not in f]
ar_files = [f for f in glob.glob(ar_pattern) if "2022-余额" not in f]

print(f"   找到 {len(ap_files)} 个 AP 流水文件")
for file in ap_files:
    try:
        df = pd.read_excel(file)
        temp = df.iloc[:, [12, 3, 16, 18]].copy()
        temp.columns = ['交易对象名称', '过账日期', '金额', '文本描述']
        temp['类型'] = 'AP'
        all_data.append(temp)
    except Exception as e:
        print(f"      ❌ AP 文件失败 {file}: {e}")

print(f"   找到 {len(ar_files)} 个 AR 流水文件")
for file in ar_files:
    try:
        df = pd.read_excel(file)
        temp = df.iloc[:, [14, 6, 20, 21]].copy()
        temp.columns = ['交易对象名称', '过账日期', '金额', '文本描述']
        temp['类型'] = 'AR'
        all_data.append(temp)
    except Exception as e:
        print(f"      ❌ AR 文件失败 {file}: {e}")

if not all_data:
    print("❌ 错误：未找到任何匹配的流水文件！")
    exit()

combined_df = pd.concat(all_data, ignore_index=True)
combined_df.to_excel(output_data_file, index=False)
print(f"💾 中间数据已保存：{output_data_file} ({len(combined_df)} 条)")

# --------------------------
# 2. 数据清洗与指标计算
# --------------------------
print("\n🧹 清洗数据、注入期初余额并计算指标...")

combined_df['过账日期'] = pd.to_datetime(combined_df['过账日期'], errors='coerce').dt.date
combined_df['日期'] = combined_df['过账日期']
combined_df['金额'] = pd.to_numeric(combined_df['金额'], errors='coerce')
combined_df = combined_df.dropna(subset=['日期', '金额', '交易对象名称'])

combined_df['应收_出货'] = 0.0
combined_df['应收_收款'] = 0.0
combined_df['应付_收货'] = 0.0
combined_df['应付_付款'] = 0.0

mask_ar = combined_df['类型'] == 'AR'
combined_df.loc[mask_ar & (combined_df['金额'] > 0), '应收_出货'] = combined_df['金额']
combined_df.loc[mask_ar & (combined_df['金额'] < 0), '应收_收款'] = -combined_df['金额']

mask_ap = combined_df['类型'] == 'AP'
combined_df.loc[mask_ap & (combined_df['金额'] < 0), '应付_收货'] = -combined_df['金额'] 
combined_df.loc[mask_ap & (combined_df['金额'] > 0), '应付_付款'] = combined_df['金额']

# 构造期初余额行
opening_rows = []
for company, amount in opening_ap_dict.items():
    if amount != 0:
        opening_rows.append({
            '交易对象名称': company,
            '日期': datetime.date(2022, 12, 31),
            '类型': 'AP_Opening',
            '应付_收货': abs(amount),
            '应收_出货': 0, '应收_收款': 0, '应付_付款': 0
        })

for company, amount in opening_ar_dict.items():
    if amount != 0:
        opening_rows.append({
            '交易对象名称': company,
            '日期': datetime.date(2022, 12, 31),
            '类型': 'AR_Opening',
            '应收_出货': abs(amount),
            '应收_收款': 0, '应付_收货': 0, '应付_付款': 0
        })

if opening_rows:
    df_opening = pd.DataFrame(opening_rows)
    combined_df = pd.concat([combined_df, df_opening], ignore_index=True)
    print(f"   ✅ 已注入 {len(opening_rows)} 条期初余额记录 (2022-12-31)")

combined_df = combined_df.sort_values(by=['交易对象名称', '日期']).reset_index(drop=True)

# --------------------------
# 3. 筛选公司
# --------------------------
print("\n📊 筛选占比 > 5% 的公司...")

cutoff_date = datetime.date(2022, 12, 31)
flow_df = combined_df[
    (combined_df['类型'].isin(['AP_Opening', 'AR_Opening'])) | 
    (combined_df['日期'] > cutoff_date)
].copy()

real_flow_df = flow_df[~flow_df['类型'].isin(['AP_Opening', 'AR_Opening'])].copy()

company_stats = real_flow_df.groupby('交易对象名称').agg(
    total_ar_collect=('应收_收款', 'sum'),
    total_ap_pay=('应付_付款', 'sum')
).reset_index()

company_stats['ratio_check'] = company_stats.apply(
    lambda x: x['total_ap_pay'] / x['total_ar_collect'] if x['total_ar_collect'] > 0 else 0, 
    axis=1
)

qualified_df = company_stats[company_stats['ratio_check'] > threshold_ratio].copy()

if qualified_df.empty:
    print("⚠️ 无符合条件的公司。")
    with pd.ExcelWriter(output_report_file, engine='openpyxl') as writer:
        pd.DataFrame({'提示': ['无公司满足条件']}).to_excel(writer, index=False, sheet_name='说明')
    exit()

print(f"   符合条件公司数：{len(qualified_df)}")

df_filtered = flow_df[flow_df['交易对象名称'].isin(qualified_df['交易对象名称'])].copy()
company_interest_map = {}

print("   预计算各公司总利息（含期末补算）...")
for company in qualified_df['交易对象名称']:
    df_comp = df_filtered[df_filtered['交易对象名称'] == company].copy()
    if df_comp.empty:
        company_interest_map[company] = 0.0
        continue
    
    daily_stats = df_comp.groupby('日期').agg(
        当天应收出货=('应收_出货', 'sum'),
        当天应收收款=('应收_收款', 'sum'),
        当天应付收货=('应付_收货', 'sum'),
        当天应付付款=('应付_付款', 'sum')
    ).reset_index()
    
    daily_stats['应收余额'] = daily_stats['当天应收出货'].cumsum() - daily_stats['当天应收收款'].cumsum()
    daily_stats['应付余额'] = daily_stats['当天应付收货'].cumsum() - daily_stats['当天应付付款'].cumsum()
    daily_stats['资金占用额'] = daily_stats['应收余额'] - daily_stats['应付余额']
    
    daily_stats['日期_dt'] = pd.to_datetime(daily_stats['日期'])
    daily_stats['间隔天数'] = daily_stats['日期_dt'].diff().dt.days.fillna(0)
    daily_stats['上期资金占用额'] = daily_stats['资金占用额'].shift(1).fillna(0)
    daily_stats['区间资金占用利息'] = daily_stats['上期资金占用额'] * annual_rate * daily_stats['间隔天数'] / 365
    
    total_interest = daily_stats['区间资金占用利息'].sum()
    
    # 【新增】计算从最后一天到今天之间的利息
    last_date = daily_stats['日期'].max()
    if isinstance(last_date, datetime.datetime):
        last_date = last_date.date()
    
    if last_date < today:
        days_to_today = (today - last_date).days
        last_occupation = daily_stats['资金占用额'].iloc[-1]
        extra_interest = last_occupation * annual_rate * days_to_today / 365
        total_interest += extra_interest
        print(f"      📈 {company}: 补算 {days_to_today} 天期末利息 = {extra_interest:.2f}")
    
    company_interest_map[company] = total_interest

qualified_df['总资金占用利息'] = qualified_df['交易对象名称'].map(company_interest_map)

summary_df = qualified_df.rename(columns={
    '交易对象名称': '公司代码',
    'total_ap_pay': '累计应付付款',
    'total_ar_collect': '累计应收收款',
    '总资金占用利息': '总资金占用利息'
})[['公司代码', '累计应付付款', '累计应收收款', '总资金占用利息']]

summary_df['sort_key'] = summary_df.apply(
    lambda x: x['总资金占用利息'] / x['累计应收收款'] if x['累计应收收款'] != 0 else 0,
    axis=1
)
summary_df = summary_df.sort_values(by='sort_key', ascending=False).reset_index(drop=True)
sorted_companies = summary_df['公司代码'].tolist()

# --------------------------
# 4. 写入 Excel
# --------------------------
print(f"\n📑 生成明细 Sheet...")

with pd.ExcelWriter(output_report_file, engine='openpyxl') as writer:
    for company in sorted_companies:
        df_comp = df_filtered[df_filtered['交易对象名称'] == company].copy()
        
        daily_stats = df_comp.groupby('日期').agg(
            当天应收出货=('应收_出货', 'sum'),
            当天应收收款=('应收_收款', 'sum'),
            当天应付收货=('应付_收货', 'sum'),
            当天应付付款=('应付_付款', 'sum')
        ).reset_index()
        
        if daily_stats.empty: continue

        daily_stats['应收余额'] = daily_stats['当天应收出货'].cumsum() - daily_stats['当天应收收款'].cumsum()
        daily_stats['应付余额'] = daily_stats['当天应付收货'].cumsum() - daily_stats['当天应付付款'].cumsum()
        daily_stats['资金占用额'] = daily_stats['应收余额'] - daily_stats['应付余额']
        
        daily_stats['日期_dt'] = pd.to_datetime(daily_stats['日期'])
        daily_stats['间隔天数'] = daily_stats['日期_dt'].diff().dt.days.fillna(0)
        daily_stats['上期资金占用额'] = daily_stats['资金占用额'].shift(1).fillna(0)
        daily_stats['区间资金占用利息'] = daily_stats['上期资金占用额'] * annual_rate * daily_stats['间隔天数'] / 365
        
        final_df = daily_stats[['日期', '应收余额', '当天应收出货', '当天应收收款', 
                                '应付余额', '当天应付收货', '当天应付付款', 
                                '资金占用额', '间隔天数', '区间资金占用利息']].copy()
        final_df.insert(0, '交易对象名称', company)
        final_df.rename(columns={
            '资金占用额': '资金占用额 (应收 - 应付)',
            '区间资金占用利息': '资金占用利息 (基于上期余额×间隔天数)'
        }, inplace=True)
        
        total_interest = final_df['资金占用利息 (基于上期余额×间隔天数)'].sum()
        total_revenue = final_df['当天应收收款'].sum()
        
        # 【新增】添加期末截止行（仅当有补算利息时）
        last_date = final_df['日期'].max()
        if isinstance(last_date, datetime.datetime):
            last_date = last_date.date()
        
        if last_date < today:
            days_to_today = (today - last_date).days
            last_occupation = final_df['资金占用额 (应收 - 应付)'].iloc[-1]
            extra_interest = last_occupation * annual_rate * days_to_today / 365
            
            end_row = {
                '交易对象名称': '期末截止',
                '日期': today,
                '应收余额': None,
                '当天应收出货': None,
                '当天应收收款': None,
                '应付余额': None,
                '当天应付收货': None,
                '当天应付付款': None,
                '资金占用额 (应收 - 应付)': last_occupation,
                '间隔天数': days_to_today,
                '资金占用利息 (基于上期余额×间隔天数)': extra_interest
            }
            final_df = pd.concat([final_df, pd.DataFrame([end_row])], ignore_index=True)
            total_interest += extra_interest
        
        cols_order = final_df.columns.tolist()
        summary_dict = {col: None for col in cols_order}
        summary_dict['交易对象名称'] = '汇总'
        summary_dict['当天应收收款'] = total_revenue
        summary_dict['资金占用利息 (基于上期余额×间隔天数)'] = total_interest
        summary_row = pd.DataFrame([summary_dict])
        final_df = pd.concat([summary_row, final_df], ignore_index=True)
        
        safe_name = str(company)[:31].replace('\\','').replace('/','').replace('?','').replace('*','').replace('[','').replace(']','')
        final_df.to_excel(writer, index=False, float_format="%.2f", sheet_name=safe_name)

final_summary_df = summary_df.drop(columns=['sort_key'])
with pd.ExcelWriter(output_report_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    final_summary_df.to_excel(writer, index=False, float_format="%.2f", sheet_name='ZZZ_Summary')

# --------------------------
# 5. 格式化 + 超链接 + 自适应列宽
# --------------------------
print("\n🔗 添加超链接、调整格式与自适应列宽...")
wb = load_workbook(output_report_file)

fill_dark_blue = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
fill_olive = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
fill_blue = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
fill_nav = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
fill_opening = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
fill_end = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # 期末行红色高亮

font_black_bold = Font(color="000000", bold=True)
font_black = Font(color="000000")
font_nav = Font(color="000000", bold=True, size=12)
font_hyperlink = Font(color="0000FF", underline="single") 
align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

def reset_cell_style(cell):
    cell.font = font_black
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.alignment = align_center

def get_text_length(value):
    if value is None:
        return 0
    s = str(value)
    if isinstance(value, (datetime.datetime, datetime.date)):
        s = value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        base_len = len(s.replace('.', '').replace('-', ''))
        base_len += 3 
        commas = (base_len - 3) // 3
        base_len += commas
        return base_len
    return len(s)

def auto_adjust_column_width(ws, min_width=15, max_width=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            length = get_text_length(cell.value)
            if length > max_len:
                max_len = length
        final_width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = final_width

def format_detail_sheet(ws, company_name):
    ws.insert_rows(1)
    cell_a1 = ws['A1']
    cell_a1.value = f"🔙 返回总汇总 (当前：{company_name})"
    cell_a1.hyperlink = "#'总汇总'!A1"
    reset_cell_style(cell_a1)
    cell_a1.font = font_nav
    cell_a1.fill = fill_nav
    cell_a1.alignment = align_center
    
    max_col = ws.max_column
    if max_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    
    headers = [cell.value for cell in ws[2]]
    if not headers: return
    
    number_cols = ['应收余额', '当天应收出货', '当天应收收款', '应付余额', '当天应付收货', '当天应付付款', '资金占用额 (应收 - 应付)', '资金占用利息 (基于上期余额×间隔天数)']
    dark_blue_cols = ['交易对象名称', '日期', '资金占用额 (应收 - 应付)', '间隔天数', '资金占用利息 (基于上期余额×间隔天数)']
    olive_cols = ['应收余额', '当天应收出货', '当天应收收款']
    blue_cols = ['应付余额', '当天应付收货', '当天应付付款']
    
    max_row = ws.max_row
    date_col_idx = headers.index('日期') + 1 
    
    for row_idx in range(2, max_row + 1):
        is_header = (row_idx == 2)
        is_sum = (not is_header) and (ws.cell(row=row_idx, column=1).value == '汇总')
        is_end = (not is_header and not is_sum) and (ws.cell(row=row_idx, column=1).value == '期末截止')
        
        is_opening = False
        if not is_header and not is_sum and not is_end:
            date_val = ws.cell(row=row_idx, column=date_col_idx).value
            if isinstance(date_val, datetime.datetime) or isinstance(date_val, datetime.date):
                if date_val.year == 2022 and date_val.month == 12 and date_val.day == 31:
                    is_opening = True
        
        fnt = font_black_bold if is_header else (font_black_bold if (is_sum or is_opening or is_end) else font_black)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None and not is_header: continue
            
            cell.font = fnt
            
            if is_header:
                cell.fill = fill_dark_blue
                cell.alignment = align_center
            else:
                if h in number_cols:
                    cell.alignment = align_right
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                elif h == '日期':
                    cell.alignment = align_center
                else:
                    cell.alignment = align_center
                
                if not is_sum:
                    if is_end:
                        cell.fill = fill_end  # 期末行红色高亮
                    elif is_opening:
                        cell.fill = fill_opening
                    elif h in dark_blue_cols: cell.fill = fill_dark_blue
                    elif h in olive_cols: cell.fill = fill_olive
                    elif h in blue_cols: cell.fill = fill_blue

    auto_adjust_column_width(ws, min_width=15, max_width=60)

def format_summary_sheet(ws, sorted_companies_list):
    headers = [cell.value for cell in ws[1]]
    if len(headers) < 4: return
    
    max_row = ws.max_row
    for col_idx in range(1, 6):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_dark_blue
        cell.font = font_black_bold 
        cell.alignment = align_center
    
    ws.cell(row=1, column=5).value = "利息占应收比重"
    
    for row_idx in range(2, max_row + 1):
        company_name = ws.cell(row=row_idx, column=1).value
        if company_name is None: continue
        
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_black
            cell.alignment = align_center if col_idx == 1 else align_right
            if col_idx > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        
        safe_sheet_name = str(company_name)[:31].replace('\\','').replace('/','').replace('?','').replace('*','').replace('[','').replace(']','')
        cell_a = ws.cell(row=row_idx, column=1)
        cell_a.hyperlink = f"#'{safe_sheet_name}'!A3"
        cell_a.font = font_hyperlink 
        
        cell_e = ws.cell(row=row_idx, column=5)
        cell_e.value = f"=IF(C{row_idx}=0, 0, D{row_idx}/C{row_idx})"
        cell_e.number_format = '0.0000%'
        cell_e.alignment = align_right

    auto_adjust_column_width(ws, min_width=15, max_width=60)

# 处理总汇总
summary_ws = wb['ZZZ_Summary']
summary_ws.title = '总汇总'
format_summary_sheet(summary_ws, sorted_companies)
wb._sheets.remove(summary_ws)
wb._sheets.insert(0, summary_ws)

# 处理明细并排序
current_sheets = [s for s in wb.sheetnames if s != '总汇总']
sheet_map = {}
for s_name in current_sheets:
    for comp in sorted_companies:
        if s_name == comp[:31]:
            sheet_map[comp] = wb[s_name]
            break

new_sheets_order = [wb['总汇总']]
for comp in sorted_companies:
    if comp in sheet_map:
        ws = sheet_map[comp]
        format_detail_sheet(ws, comp)
        new_sheets_order.append(ws)
    else:
        for s in current_sheets:
            if s not in [sh.title for sh in new_sheets_order]:
                ws = wb[s]
                new_sheets_order.append(ws)

wb._sheets = new_sheets_order
wb.save(output_report_file)

print("="*80)
print("✅ 全部完成！")
print(f"📁 文件路径：{os.path.abspath(output_report_file)}")
print("✨ 本次更新：")
print("   1. 已自动补算从‘最后一笔交易’到‘今天’的利息")
print("   2. 明细表中新增‘期末截止’行（红色高亮），显示补算天数和利息")
print("   3. 总表中的‘总资金占用利息’已包含期末补算部分")
print("   4. 所有原有功能（自适应列宽、黑字、正数应付等）保持不变")
print("="*80)